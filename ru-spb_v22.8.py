import mimetypes
import os
import smtplib
import time
from datetime import date
from datetime import datetime
from datetime import timedelta
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from openpyxl.styles import PatternFill
import investpy
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from pandas import DataFrame
from sqlalchemy import create_engine
from tqdm import tqdm
from threading import Thread
import threading

from pr_config import *  ## пробуем подгрузить свои данные
from sec_config import *

current_version = "22.8- sql"

'''
делаем массив для анализа ошибок!!

v22.2 - изменения в связи с заменой источника branch - теперь не из файла, а из базы данных, но там пока нет для US почему то..
!
v22.3 - изменения - оптимизируем код  convert_df_to_np_from_sql - сокращаем число строк и переменные подгружаем в списки
        добавил окраску ячеек в экселе по списку из D_list. 
v22.4 - изменения... добавляем модуль мониторинга.... возможно удалим very_vol, с-200, 

v22.6 - пробуем добавить мультипоточность, кстати - 1--- в процесс подгрузки из sql. 2 ---  в расчеты 
        и поменять способ добавления строки в датафрейм - делаем через df.loc[len(df)] = [**kwarg] -- выделено в отдельную ветку 
        вводим потоки через словарь - именованные обращения получаются а не порядковые!!!!
        
         --- требуется - написать программу, которая будет по списку... (список вручную) .. из базы данных вырезать все значения тикера 
        у которого прошел сплит..  и загружать новые данные за всю историю  - написали это split_check_update .
        
        --- надо добавить pr_config.py с общими переменными для всех файлов проекта.
    
v22.7 добавлена многопоточность, изменен список получателей почты - по первому списку получают  1 файл, по второму списку - уже два
        в логах по итогам пишем время в формате Ч:М:С
        добавлен pr_config.py с общими переменными для всех файлов проекта и секретный файл с паролями от почты,
         !!! хотя sql_login тоже соделжит пароль от базы данных

v22.8   добавлена запись расчетных датафреймов в таблицу mysql --- tiker_report
        -- те же столбцы, плюс market_name .. остальное вроде такое же все 
TODO ::::
        !!!сейчас записываем по итогам расчетов , плюс наверное нужно делать таблицу статуса - потом будем проверять -- может уже расчитано все!!
        !!! или может делать налету - построчно..  
        Select tiker, max(day_close) as day_close_max, market from tiker_report group by tiker;
        !!сделать возможность добавления кучи списков тикеров - и чтоб ексель добавлял странички налету - при любом кол-ве списков
        !!! упростить шаблон екселя - все равно все страницы равны --- тока разница в таблице теханализа..  
        1000200) надо придумать шапку головную - что когда запускать --- вот модуль insert_history_date_into_sql(): - когда запускаем ??   может его в update_sql.py перенести???          
'''


# TODO: добавить модуль для считывания из базы данных mysql и в эксель!!! и потом дальнейшее форматирование)))

def bif_report_tables_to_sql(df, market):
    """записываем расчетную табличку в sql базу таблица -- tiker_report"""
    df['market'] = market
    df.set_index('tiker', inplace=True)
    try:
        df.to_sql(name='tiker_report', con=create_engine(sql_login), if_exists='append')  # append , replace
    except Exception as _ex:
        save_exeption_log(linux_path, modul='report_table', message=str(_ex))
        print(f'SQL save errorrr {str(_ex)}\n ', df.shape, df, '\n')


def my_hist_data_from_spb_list_update():
    '''подпрограмма изменения market c US на SPB в базе данных на основании ListingSecurityList.csv
    1) необходимо просто подсунуть ListingSecurityList.csv и все (проверить не поменялось ли название столбца с
    названиями тикеров... дальше все будет все само.
    '''
    save_log(prj_path, message='start update SPB list update from CSV file_____')
    db_connection = create_engine(sql_login)  # connect to database
    df_last_update = pd.read_sql('Select * from base_status ;', con=db_connection)
    df_spb = pd.read_csv('ListingSecurityList.csv', delimiter=';')
    df_spb = df_spb[df_spb['s_currency_kotir'] == 'USD']['s_RTS_code']
    save_log(prj_path, message='all data from file load --')
    print(f"spb birga {len(df_spb)}")
    spb_df_last = df_last_update[df_last_update['market'] == 'US']['st_id']
    print(f'my df {len(spb_df_last)}')
    save_log(prj_path, message=f'tiker on SPB{len(df_spb)}, tiker on my sql_base{len(spb_df_last)}')
    listNew = []
    for index in df_spb:
        if spb_df_last.isin([index]).any():
            listNew.append(index)
    print(len(listNew))
    if len(listNew) != 0:
        print(len(listNew))
        print(listNew)
        for index in listNew:
            db_connection.execute(f"update hist_data set market='SPB' where st_id = '{index}';")
            print(index)
        print(f"UPdate Complite. {len(listNew)} st_id market from [US] to [SPB]")
        save_log(prj_path, message=f"UPdate Complite. {len(listNew)} st_id market from [US] to [SPB]")
        save_log(prj_path, message=f"writen next tiker {listNew}")


def insert_history_date_into_sql():
    """
    загружаем новые тикеры в sql базу на основе investpy.get_stocks.
    вставляем под market US

    ?? а какую дату загрузки делать??? - начальную можно сделать 1/1/2019, а конец - дату сегодняшнюю???
    :return:
    """
    db_connection = create_engine(sql_login)
    market_name = ['United States', 'United States', 'russia']
    stocks_us_investpy = investpy.get_stocks(country=market_name[0])['symbol']
    df_last_update = pd.read_sql('Select * from base_status ;', con=db_connection)
    # df_last_update = pd.read_excel('base_status.xlsx')['st_id']
    save_log(prj_path, message='insert new tiker from investpy.get_stocks to SQL history_date')
    df = df_last_update.to_numpy().tolist()
    my_2_list = stocks_us_investpy.to_numpy().tolist()
    my_only_US_df_list = []
    for index_us in tqdm(my_2_list):
        # print(index_us)
        if index_us in df:
            # print(f'is in -={index_us}')
            continue
        else:
            my_only_US_df_list.append(index_us)
            # print(f"{index_us} - add" )

    print(f"my list len = {len(my_only_US_df_list)}")
    print(f'all US list len = {len(my_2_list)}')
    my_only_US_df_list.sort()
    save_log(prj_path, message=f'find {len(my_only_US_df_list)}, try add {my_only_US_df_list}')
    print(my_only_US_df_list)
    from_date_m, to_date_m = '4/02/2020', '6/02/2020'
    successfully_list = []
    for only_us_index in tqdm(my_only_US_df_list):
        try:
            df_update = investpy.get_stock_historical_data(stock=only_us_index,
                                                           country=market_name[0],
                                                           from_date=from_date_m,
                                                           to_date=to_date_m)
            time.sleep(1.0)
            df_update[['st_id', 'Currency', 'market']] = only_us_index, "USD", 'US'
            successfully_list.append(only_us_index)
        except:
            print(f'Error [{only_us_index}] loading')
            continue
        pd_df_to_sql(df_update)
    save_log(prj_path,
             message=f'LEn of successfully list is [{len(successfully_list)}], apply [{round(100 * len(successfully_list) / len(my_only_US_df_list), 0)}]%,  added list-- {successfully_list}')


def stock_name_table(prj_path):
    """ старое....
    !!!! не  используется"""
    ru_stos1 = investpy.stocks.get_stocks(country='russia')
    ru_stos = ru_stos1[['name', 'symbol']]
    mmm = np.zeros((len(ru_stos), 1)) + 3
    my_id = pd.DataFrame(data=mmm, columns=['markets_id'])
    ru_stos = ru_stos.join(my_id)
    # print(ru_stos)
    big_df = pd.DataFrame(columns=['name', 'symbol'])
    big_df = big_df.append(other=ru_stos)
    df_spb = pd.read_csv('ListingSecurityList.csv', delimiter=';',
                         encoding='cp1251')  # читаем список всех тикеров на СПБ
    stock_spb = df_spb[df_spb['s_currency_kotir'] == 'USD']  # вычисляем список эмитентов СПБ с курсом в USD
    us_st_spb = stock_spb[['name', 'symbol']]  # срезаем лишние столбцы из списка -- надо менять название столбца!!!!!
    mmm = np.zeros((len(df_spb), 1)) + 2
    my_id = pd.DataFrame(data=mmm, columns=['markets_id'])
    us_st_spb = us_st_spb.join(my_id)
    # print(us_st_spb)
    us_stos_usa1 = investpy.stocks.get_stocks(country='United States')
    us_stos_usa = us_stos_usa1[['name', 'symbol']]

    mmm = np.zeros((len(us_stos_usa), 1)) + 1
    my_id = pd.DataFrame(data=mmm, columns=['markets_id'])
    us_stos_usa = us_stos_usa.join(my_id)
    # print(us_stos_usa)
    big_df = big_df.append(other=us_st_spb, ignore_index=True)
    big_df = big_df.append(other=us_stos_usa, ignore_index=True)
    # print('all', big_df)

    with pd.ExcelWriter(prj_path + 'my_all_st.xlsx') as writer:
        big_df.to_excel(writer, sheet_name='all')  ### работает!!!
    # big_df.to_csv('my_test_all_st.csv', sep=';', encoding='cp1251', line_terminator='/n', index=True)

    exit()


def year_live_stock():  # df, stick_id, st_name): #
    ''' расчет годового роста каждый месяц --
    еще не доделано!!!!! '''
    print('Блок расчета годоводго роста по месяцам')
    df_col_list = ['name', 'stiker', 'today_close', 'inter1', 'inter2', 'inter3', 'inter4', 'inter5', 'inter6',
                   'inter7', 'inter8', 'inter9', 'inter10', 'inter11', 'inter12', 'inter_sum', 'inter_mean']
    '''' inter1 - интервал с даты отчета - на год назад, inter12 -- интервал с года назад до 2 лет назад, остальные промежуточные 
    за базу расчетов берем среднее из максимумов и минимумов за десять дней до даты отсчета'''
    df_date = pd.DataFrame(
        columns=['date_start_now', 'date_end_now', 'date_start_12', 'date_end_12'])  # создали датафрейм со всеми датами
    stick_id, st_name = 'AAPL', 'Apple'

    df_from_sql = pd.read_sql(
        'Select date, high, low, close, st_id, Currency from hist_data  WHERE Currency=\'USD\' and st_id = \'AAPL\';',
        con=sql_login)  ## загружаем базу СПБ ---USD
    k, y_q, m_d, m_m = 0, 0, 0, 0
    cur_date = datetime.today()
    for i in range(0, 12):
        if date.today().month - i + k <= 0:
            k, y_q = 12, 1
        if date.today().day - 10 < 0:
            m_d, m_m = 29, -1
        df_date.loc[i, 'date_start_now'] = date(date.today().year - y_q, date.today().month - i + k + m_m,
                                                date.today().day - 10 + m_d)
        df_date.loc[i, 'date_end_now'] = date(date.today().year - y_q, date.today().month - i + k, date.today().day)
        df_date.loc[i, 'date_start_12'] = date(date.today().year - y_q - 1, date.today().month - i + k + m_m,
                                               date.today().day - 10 + m_d)
        df_date.loc[i, 'date_end_12'] = date(date.today().year - y_q - 1, date.today().month - i + k, date.today().day)
    # print(df_date.tail(3))
    # print(df_from_sql.tail(10))
    delta_np = np.zeros(12)
    print(delta_np)
    indexx = 0
    for indexx in range(len(df_date)):
        # print(df_date.date_start_now[indexx],df_date.date_end_now [indexx] )
        mean_delta = 0.5 * (pd.DataFrame.mean(
            df_from_sql[df_from_sql.date > pd.to_datetime(df_date.date_start_now[indexx], format='%Y-%m-%d')].iloc[:10][
                'high'])
                            + pd.DataFrame.mean(df_from_sql[
                                                    df_from_sql.date > pd.to_datetime(df_date.date_start_now[indexx],
                                                                                      format='%Y-%m-%d')].iloc[:10][
                                                    'low']))
        mean_delta2 = 0.5 * (pd.DataFrame.mean(
            df_from_sql[df_from_sql.date > pd.to_datetime(df_date.date_start_12[indexx], format='%Y-%m-%d')].iloc[:10][
                'high'])
                             + pd.DataFrame.mean(
                    df_from_sql[
                        df_from_sql.date > pd.to_datetime(df_date.date_start_12[indexx], format='%Y-%m-%d')].iloc[
                    :10]['low']))
        delta_np[indexx] = round((mean_delta - mean_delta2) / mean_delta2, 3)
        # print('d1',round(mean_delta,2),'d2', round(mean_delta2,2), 'delta_pr', delta_np[indexx])
    print('delta', delta_np)
    print('mean', round(delta_np.mean(), 3))
    print('sum', delta_np.sum())

    year_live_stock_df = pd.DataFrame(columns=df_col_list)
    #    year_live_stock_df = year_live_stock_df[]

    exit()


def save_history(prj_path, big_df_table):
    big_df_table.to_excel(prj_path + 'history_data.xlsx')


def save_log(prj_path, message):  # сохраняет в лог файл сообщение..
    f = open(prj_path + 'make_from_sql.log', mode='a')
    lines = '[' + str(datetime.today()) + '] ' + str(message + ' v' + current_version)
    f.writelines(lines + '\n')
    f.close()
    # print(lines+ '\n')


def input_tables_append(sql_login, prj_path):
    st_name = pd.read_excel(prj_path + 'my_all_st.xlsx', index_col=0)
    branch_name = pd.read_excel(prj_path + 'tiker-branch.xlsx', index_col=0)
    # db_connection = create_engine(sql_login)
    # df_last_update = pd.read_sql('Select st_id, max(date) as date_max, Currency  from hist_data group by st_id',                                 con=db_connection)
    col_list_m = ['name', 'tiker', 'branch', 'currency', 'market', 'max_date', 'min_date', 'status_table']
    print(st_name.head(3), len(st_name))

    print(branch_name.head(3), len(branch_name))
    # print(df_last_update.head(3), len(df_last_update))
    status_t = pd.DataFrame(columns=col_list_m)
    print(status_t)
    status_index = 0
    for ind in range(len(branch_name.st_id)):
        print(branch_name['st_id'].iloc[ind])
        if pd.DataFrame.any(status_t['st_id'] != branch_name['st_id'].iloc[ind]):
            status_t['st_id'].iat[status_index] = branch_name['st_id'].iloc[ind]
            status_index += 1

    print(status_t, status_index)


def convert_df_to_np_from_sql(df1, stok_name, st_inv, col_list, branch,
                              df_teh):  # подчтет данных их исторических значений
    ''' оптимизировал код - исключен нумпай и прочие строчки,
    округлил до целых проценты
    '''
    time_list = [-1, -5, -10, -15, -20, -30, -40, -50, -60]
    today_close = df1.iloc[-1]['close']
    day_close = df1.iloc[-1]['date'].date()
    day_start = df1.iloc[0]['date'].date()
    # df1.index[0].date()
    min_y = df1.iloc[:]['low'].min()  # + 0.00000001 # иногда минимум =0, поэтому добавляем мизер - чтоб не делить на 0
    max_y = df1.iloc[:]['high'].max()
    # min_y_date = df1[df1['low'] == min_y].iloc[0]['date'].date()
    # max_y_date = df1[df1['high'] == max_y].iloc[0]['date'].date()
    if df1.iloc[-10:-1]['close'].min() < 3.0:
        my_round = 6
    else:
        my_round = 2
    min_dek, max_dek = [], []
    min_per, max_per = [], []
    min_pr_delta, max_pr_delta = [], []
    for index in range(len(time_list) - 1):
        min_dek.append((df1.iloc[time_list[index + 1]:time_list[index]]['low'].min()).round(my_round))
        max_dek.append((df1.iloc[time_list[index + 1]:time_list[index]]['high'].max()).round(my_round))
    for index_2 in range(0, 8):
        min_per.append(round((today_close / (min_dek[index_2] + 0.000001) - 1) * 100, 0))
        max_per.append(round((today_close / max_dek[index_2] - 1) * 100, 0))
    for index in range(0, 7):
        min_pr_delta.append(round(min_per[index] - min_per[index + 1], 0))
        max_pr_delta.append(round(max_per[index] - max_per[index + 1], 0))
        # print( df1.iloc[time_list[index + 1]:time_list[index]]['low'].min(), index, time_list[index], time_list[index + 1])
    today_y_pr_max = round((today_close / max_y - 1) * 100, 0)
    today_y_pr_min = round((today_close / (min_y + .000001) - 1) * 100, 0)
    m1_max, m1_min = df1.iloc[-20:-1]['high'].max(), df1.iloc[-20:-1]['low'].min()
    m3_max, m3_min = df1.iloc[-60:-1]['high'].max(), df1.iloc[-60:-1]['low'].min()
    m6_max, m6_min = df1.iloc[-120:-1]['high'].max(), df1.iloc[-120:-1]['low'].min()
    year1_max, year1_min = df1.iloc[-240:-1]['high'].max(), df1.iloc[-240:-1]['low'].min()
    pr_30_day_max, pr_30_day_min = round((today_close / m1_max - 1) * 100, 0), round(
        (today_close / (m1_min + 0.000001) - 1) * 100,
        0)
    pr_90_day_max, pr_90_day_min = round((today_close / m3_max - 1) * 100, 0), round(
        (today_close / (m3_min + 0.0000001) - 1) * 100,
        0)
    pr_6_m_max, pr_6_m_min = round((today_close / m6_max - 1) * 100, 0), round(
        (today_close / (m6_min + .0000001) - 1) * 100, 0)
    pr_1y_max, pr_1y_min = round((today_close / year1_max - 1) * 100, 0), round(
        (today_close / (year1_min + .00000001) - 1) * 100, 0)
    my_frame = [st_inv, stok_name, branch, today_close, *min_dek, *min_per, *min_pr_delta,
                *max_dek, *max_per, *max_pr_delta,
                min_y, max_y, today_y_pr_min, today_y_pr_max, day_start, day_close, *df_teh,
                m1_max, m1_min, m3_max, m3_min, m6_max, m6_min, year1_max, year1_min, pr_30_day_max, pr_30_day_min,
                pr_90_day_max, pr_90_day_min,
                pr_6_m_max, pr_6_m_min, pr_1y_max, pr_1y_min]
    # , min_date, max_date]
    df2 = pd.DataFrame([my_frame], columns=list(col_list))
    # print('DF___',df2)
    return df2  # готовим данные для записи в таблицу


def statistic_data_base(df_last_update):
    global prj_path
    ''' подсчет статистики актуальности базы данных, с охранением результата в лог'''
    for market_s in df_last_update['market'].unique():
        listing_ll = pd.Series(
            {c: df_last_update[df_last_update['market'] == market_s][c].unique() for c in df_last_update})
        listing_ll['date_max'].sort()

        for num_1 in range(len(listing_ll[2]) - 2, len(listing_ll[2])):
            print(
                f"date for [{market_s}]- [{str(pd.to_datetime(listing_ll['date_max'][num_1]).date())}] is [{len(df_last_update[(df_last_update['market'] == market_s) & (df_last_update['date_max'] == listing_ll['date_max'][num_1])]['date_max'])}] ")
            save_log(prj_path,
                     f"date for [{market_s}]=[{str(pd.to_datetime(listing_ll['date_max'][num_1]).date())}] is [{len(df_last_update[(df_last_update['market'] == market_s) & (df_last_update['date_max'] == listing_ll['date_max'][num_1])]['date_max'])}] ")


def sql_base_make(prj_path, sql_login,
                  col_list):  # Модуль загрузки данных из базы mysql и формирования отчетных таблиц
    global branch_name_local
    start_timer = datetime.today()
    db_connection = create_engine(sql_login, connect_args={'connect_timeout': 10})  # connect to database
    big_df: DataFrame = pd.DataFrame(columns=list(col_list))
    big_df_US: DataFrame = pd.DataFrame(columns=list(col_list))
    today_date = datetime.today()
    teh_an_list = ['teh_daily_sel', 'teh_daily_buy', 'teh_weekly_sel', 'teh_weekly_buy', 'teh_monthly_sell',
                   'teh_monthly_buy', 'daily_sma_signal 200', 'daily_ema_signal 200', 'weekly_sma_signal 200',
                   'weekly_ema_signal 200', 'monthly_sma_signal 200', 'monthly_ema_signal 200', 'EPS', 'P_E']
    teh_an_df_nodata = pd.DataFrame(columns=teh_an_list)
    teh_an_df_nodata.loc[len(teh_an_df_nodata)] = 'No DAta'
    big_df_ru = pd.DataFrame(columns=list(col_list))
    thre_sql_return, thread_link = {}, {}

    def thread_sql_q(key, sql_command):  # многопоточные запросы в sql  через словарь
        local_start_time = time.time()
        if key == 'hist_US':
            if datetime.today().weekday() == 5:
                try:
                    thre_sql_return[key] = pd.read_sql(sql_command, con=db_connection)
                except Exception as _ex:
                    print(_ex)
        else:
            try:
                thre_sql_return[key] = pd.read_sql(sql_command, con=db_connection)
            except Exception as _ex:
                print(f'error thread SQL {_ex}')
        delta_time = round(time.time() - local_start_time, 2)
        print('-----END load SQL Thread for ', key, f' time to complite [{delta_time}] sec')

    # команды для запросов в базу данных в многопоточном режиме
    sql_comm_key = ['tiker_branch', 'base_status', 'teh_an_status', 'hist_SPB', 'hist_RU', 'hist_US', 'teh_an_base']
    sql_command_list = ['Select * from tiker_branch ;',
                        'Select * from base_status ;',
                        'Select st_id, max(date) as date_max from teh_an group by st_id',
                        f'Select date, high, low, close, st_id, Currency from hist_data  WHERE market=\'SPB\' and date > \'{my_start_date}\';',
                        f'Select date, high, low, close, st_id, Currency from hist_data  WHERE market=\'RU\' and date > \'{my_start_date}\';',
                        f'Select date, high, low, close, st_id, Currency from hist_data  WHERE market=\'US\' and date > \'{my_start_date}\';',
                        'Select hd.* from teh_an hd join (Select hd.st_id, max(hd.date) as date_max from teh_an hd group by hd.st_id) hist_data_date_max on hist_data_date_max.st_id = hd.st_id and hist_data_date_max.date_max = hd.date;'
                        ]
    for key, index_name in zip(sql_comm_key, sql_command_list):
        thread_link[key] = threading.Thread(target=thread_sql_q, args=(key, index_name,))

    thread_link[sql_comm_key[0]].start()
    thread_link[sql_comm_key[0]].join()
    thread_link[sql_comm_key[1]].start()

    '''
    надо придумать как отлавливать содержимое thre_sql_return - данные могут записаться в другом порядке ---- 
    наверное надо по колонкам определяться - .column -- и смотреть..
    '''
    # отрасли
    branch_name = thre_sql_return[sql_comm_key[0]]  ##pd.read_sql(sql_command_list[0], con=db_connection)
    thread_link[sql_comm_key[2]].start()
    # tiker , name, branch,  curency
    ###

    print(f'D dataframe list  created [{len(dmitry_list_spb)}]')
    print(f'Z dataframe list  created [{len(zina_list_spb)}]')
    print('[', datetime.today(), ']', "Dataframe load...[OK]")
    save_log(prj_path, "Dataframe load...[OK]")

    thread_link[sql_comm_key[1]].join()

    df_last_update = thre_sql_return[sql_comm_key[1]]  ###pd.read_sql(sql_command_list[1], con=db_connection)
    last_day_sql = df_last_update.iloc[1]['date_max'].date()
    # if (today_date.date() - last_day_sql).days != 0 :
    #     df_last_update = pd.read_sql(
    #         'Select st_id, max(date) as date_max, Currency, min(date) as date_min , market from hist_data group by st_id;',
    #         con=db_connection)  # загрузили список тикеров из базы с последней датой
    #     df_last_update['today_day'] = datetime.today()
    #     df_last_update.to_sql(name='base_status', con=db_connection, if_exists='replace')  # append , replace
    #     save_log(prj_path, 'base_status table is Update ')

    thread_link[sql_comm_key[3]].start()
    thread_link[sql_comm_key[2]].join()
    thread_link[sql_comm_key[6]].start()
    df_last_teh = thre_sql_return[sql_comm_key[2]]

    '''
     считаем статистику по теханализу 
    '''
    max_teh_date = pd.DataFrame.max(df_last_teh.date_max[:])
    teh_full = round(len(df_last_teh[df_last_teh.date_max == max_teh_date]) / len(df_last_teh), 2)
    min_teh_date = pd.DataFrame.min(df_last_teh.date_max[:])
    print('teh_Full', teh_full * 100, '% Have max date')
    save_log(prj_path, 'teh_full ' + str(teh_full * 100) + '% have MAX date')
    if teh_full != 1:
        save_log(prj_path, 'base teh analis not full - only ' + str(teh_full * 100) + ' %')
        max_teh_date = min_teh_date  # берем для использования минимальную из максимальных
    save_log(prj_path, 'teh_date - [' + str(max_teh_date) + ']')
    max_stick, start_stick_num = len(df_last_update['st_id']), 0
    max_date = df_last_update.date_max[:].max()
    print('max date for hist_date', max_date)

    for indexx in df_last_update['st_id']:
        if (today_date - df_last_update[df_last_update.st_id == indexx].iloc[0]['date_max']).days <= max_old_days:
            start_stick_num += 1
            # print(indexx)
    message = f'hist_data tiker smoler [{max_old_days}] days [{start_stick_num}], all tikers [{max_stick}], part=[{round(100 * start_stick_num / max_stick, 1)}] %'
    save_log(prj_path, message)
    print(message)
    df_out_date = df_last_update[df_last_update.date_max <= max_date - timedelta(days=max_old_days)]
    df_last_update = df_last_update[df_last_update.date_max > max_date - timedelta(
        days=max_old_days)]  ### отрезаем все тикеры , для которых данные старые!!!!!!!!! --- так можно все порезать так, что считать будет нечего
    statistic_data_base(df_last_update)
    df_out_date.sort_values(by=['date_max'], inplace=True)
    save_log(prj_path,
             f'len of outdate of hist_date {len(df_out_date)} from {len(df_last_update)} id [{round(100 * len(df_out_date) / len(df_last_update), 0)}]%')

    thread_link[sql_comm_key[3]].join()
    thread_link[sql_comm_key[4]].start()
    df_spb = thre_sql_return[sql_comm_key[3]]

    print(f'SPB sql load OK [{df_spb.shape}]')

    if datetime.today().weekday() == 5:
        thread_link[sql_comm_key[5]].start()

    save_log(prj_path, "SPB base load--" + str(len(df_spb)) + "...[OK]")

    '''
    загружаем базу данных, и список тикеров с последней датой обновления . сортируем по дате. по списку
    тикеров проходимся в базе данных -
    делаем выборку по тикеру - столбцы date, high, low -- переводим индекс date. уже не конвертируем в numpy.
    передаем в функцию
    '''
    # TODO: надо заменить весь этот блок вызовом компактных функций. возможно с мультипроцессорностью.\
    #
    df_spb.sort_values(by=['date'], inplace=True)
    thread_link[sql_comm_key[6]].join()
    df_teh = thre_sql_return[sql_comm_key[6]]
    print('\n[', datetime.today(), ']', 'stage 1 (SPB).....[Calculating]')
    for ind in tqdm(df_last_update['st_id']):  # считаем рынок СПБ
        if pd.DataFrame.any(branch_name.st_id == ind):
            branch_name_local = branch_name[branch_name.st_id == ind].iloc[0]['branch']
        else:
            branch_name_local = 'NO data'
        if df_last_update[df_last_update.st_id == ind].iloc[0]['market'] == "SPB":
            df_1 = df_spb[df_spb['st_id'] == ind]
            # print(df_1)
            # print('name', st_name[st_name.symbol== ind].iloc[0]['name'])
            # print(branch_name[branch_name.st_id == ind].iloc[0]['name'])
            # print(ind, branch_name_local)
            try:
                my_df = convert_df_to_np_from_sql(df_1, branch_name[branch_name.st_id == ind].iloc[0]['name'], ind,
                                                  col_list,
                                                  branch_name_local,
                                                  df_teh[df_teh.st_id == ind].iloc[0][3:])  # пробуем добавить теханализ
            except Exception as _ex:
                save_log(prj_path, str(_ex))
                message = 'SPB error ' + str(ind)
                save_log(prj_path, message)
                continue
            # print(my_df)
            big_df = pd.concat([big_df, my_df])
    print('len big df SPB', len(big_df))

    bif_report_tables_to_sql(big_df.copy(), 'SPB')

    thread_link[sql_comm_key[4]].join()
    df_ru = thre_sql_return[sql_comm_key[4]]
    df_ru.sort_values(by=['date'], inplace=True)
    save_log(prj_path, "RU base load--" + str(len(df_ru)) + "...[OK]")
    print('\n[', datetime.today(), ']', 'stage 2 (RU)......[Calculating]')
    for ind in tqdm(df_last_update['st_id']):  # считаем рынок Московская биржа
        if (branch_name.st_id == ind).any():
            branch_name_local = branch_name[branch_name.st_id == ind].iloc[0]['branch']
        else:
            branch_name_local = 'NO data'
        if df_last_update[df_last_update.st_id == ind].iloc[0]['market'] == "RU":
            df_2 = df_ru[df_ru['st_id'] == ind]
            try:
                my_df_ru = convert_df_to_np_from_sql(df_2, branch_name[branch_name.st_id == ind].iloc[0]['name'],
                                                     ind, col_list, branch_name_local,
                                                     df_teh[df_teh.st_id == ind].iloc[0][3:], )
            except Exception as _ex:
                save_log(prj_path, str(_ex))
                message = 'RU error ' + str(ind)
                save_log(prj_path, message)
                continue
            big_df_ru = pd.concat([big_df_ru, my_df_ru])
    print('\n[', datetime.today(), ']', 'stage 3 (dmitry)..[Collecting]')
    ###  создаем сборку для фильтрации – это простая выборка из массива на основе шаблона – гораздо проще чем поиск
    # кстати надо попробовать предыдущие циклы переделать с помощью команды map
    # – наверное не получится – слишком много переменных передавать надо
    ### зато интересный скилл изучили..!!!
    dmitry_list_for_filter = big_df['tiker'].isin([*dmitry_list_spb])
    dmitry_df = big_df[dmitry_list_for_filter].copy()
    print('\n[', datetime.today(), ']', 'stage 4 (zina)..[Collecting]')
    zina_df = big_df[big_df['tiker'].isin([*zina_list_spb])].copy()

    if datetime.today().weekday() == 5:

        thread_link[sql_comm_key[5]].join()
        df_from_us = thre_sql_return[sql_comm_key[5]]

        df_from_us.sort_values(by=['date'], inplace=True)
        save_log(prj_path, "US base load--" + str(len(df_from_us)) + "...[OK]")
        print('\n[', datetime.today(), ']', 'stage 5 (US)..[Collecting]')
        for tik_index in tqdm(df_last_update['st_id']):
            if pd.DataFrame.any(branch_name.st_id == tik_index):
                branch_name_local = branch_name[branch_name.st_id == tik_index].iloc[0]['branch']
            else:
                branch_name_local = 'NO data'
            if df_last_update[df_last_update.st_id == tik_index].iloc[0]['market'] == "US":
                df_1 = df_from_us[df_from_us['st_id'] == tik_index]
                # print('name', st_name[st_name.symbol== ind].iloc[0]['name'])
                try:
                    my_df = convert_df_to_np_from_sql(df_1, branch_name[branch_name.st_id == tik_index].iloc[0]['name'],
                                                      tik_index,
                                                      col_list,
                                                      branch_name_local,
                                                      df_teh[df_teh.st_id == tik_index].iloc[0][
                                                      3:])  # пробуем добавить теханализ
                except Exception as _ex:
                    save_log(prj_path, str(_ex))
                    message = 'US error ' + str(tik_index)
                    save_log(prj_path, message)
                    continue
                big_df_US = pd.concat([big_df_US, my_df])  # , list(col_list))
    else:
        print('\n[', datetime.today(), ']', 'today not for stage 5 -- (US), see later -(need 5 day of weekday)')

    name_for_save = str(prj_path) + 'sql_make-' + str(datetime.today().date()) + '.xlsx'
    with pd.ExcelWriter(name_for_save) as writer:  # записываем отчетный файл
        big_df.to_excel(writer, sheet_name='SPB')  # считаем весь рынок SPB
        big_df_ru.to_excel(writer, sheet_name='RU')  # считаем весь рынок RU
        big_df_US.to_excel(writer, sheet_name='USA')  # считаем весь рынок USA
        df_teh.to_excel(writer, sheet_name='Teh_analis_all')  #
        dmitry_df.to_excel(writer, sheet_name='D_list')
        zina_df.to_excel(writer, sheet_name='Z_list')
        # считаем весь рынок + волатильность RU
        # считаем весь рынок + волатильность SPB

    print('\n[', datetime.today(), ']', 'EXCEL file .... [SAVED]')
    print('\n[', datetime.today(), ']', (datetime.today() - start_timer).seconds, '[sec]')
    save_log(prj_path, '-----[' + str(datetime.today()) + '] ' + str(
        timedelta(seconds=((datetime.today() - start_timer).seconds))))
    # bif_report_tables_to_sql(big_df, 'SPB')
    bif_report_tables_to_sql(big_df_ru.copy(), 'RU')
    bif_report_tables_to_sql(big_df_US.copy(), 'USA')
    save_log(prj_path, 'SQL save complite')
    # exit()
    return name_for_save


def excel_format(prj_path, name_for_save, history_path):  # форматирование отчетного файла после записи
    redFill = PatternFill(start_color='FFFF0000',
                          end_color='FFFF0000',
                          fill_type='solid')

    wb_help = load_workbook(filename=str(prj_path) + xlsx_sample)  # шаблон форматирования
    shit_n = wb_help.sheetnames
    ws = wb_help.active
    print('[', datetime.today(), ']', 'EXCEL format [START]')
    print(shit_n)
    d_list_list = []
    for shit_m in tqdm(shit_n):
        ws = wb_help[shit_m]
        df_l = pd.read_excel(name_for_save, engine='openpyxl', sheet_name=shit_m)
        print(ws)
        for r in dataframe_to_rows(df_l, index=False, header=False):
            ws.append(r)
        ws.auto_filter.ref = "A6:" + get_column_letter(ws.max_column) \
                             + str(ws.max_row)
        if shit_m == 'D_list':
            for index_x in df_l['tiker']:
                d_list_list.append(index_x)
        if shit_m == 'SPB':
            df_l_spb = df_l.copy()
    if len(d_list_list) != 0:
        my_filter = df_l_spb['tiker'].isin([*d_list_list])
        ws = wb_help['SPB']
        for index_x in df_l_spb[my_filter].index:
            ws[f'B{index_x + 7}'].fill = redFill
        print(f"SPB tiker from d_list -- fill [{len(d_list_list)}]")

        # ws.freeze_panes = 'E7'
    # print(f'Now remove old file [{os.pardir} {name_for_save}]')
    # save_log(prj_path, str(f'Now remove old file [{os.pardir} {name_for_save}]'))
    # os.remove(name_for_save)
    for s_name in shit_n:
        sheet = wb_help[s_name]
        m_row = sheet.max_row
        m_col = sheet.max_column
        print(s_name, m_row, m_col)

    cur_date_for_file_name = str(date.today().day) + '-' + str(date.today().month) + '-' + str(date.today().year)
    name_for_save = str(prj_path) + "d-отчет-" + cur_date_for_file_name + ".xlsx"
    wb_help.save(name_for_save)
    save_log(prj_path, str(wb_help.sheetnames))
    wb_help.save(str(history_path) + "d-отчет-" + cur_date_for_file_name + ".xlsx")  # save to Yandex drive
    pfd = wb_help['D_list']
    wb_help.remove(pfd)
    pfd = wb_help["Z_list"]
    wb_help.remove(pfd)
    save_log(prj_path, str(wb_help.sheetnames))
    name_for_save_crop = str(history_path) + "all-отчет-" + cur_date_for_file_name + ".xlsx"
    wb_help.save(name_for_save_crop)

    print('[', datetime.today(), ']', 'EXCEL format.. [Complite]')
    print('[', datetime.today(), ']', 'EXCEL save to YD.. [Complite]')

    return name_for_save, name_for_save_crop


def pd_df_to_sql(df, sql_login, st_name, ind,
                 len_max):  # запись исторических значений при update с investpy  в базу MYSQL
    engine = create_engine(sql_login)
    try:
        df.to_sql(name='hist_data', con=engine, if_exists='append')  # append , replace
    except:
        print('pd_to_sql --- [error]')


def send_email(name_for_save, name_for_save_crop):
    """ пробуем отправлять почту
    для отправки отбираются два файла - name_for_save_crop, name_for_save, склеиваются в один пакет 
    и отправляется по списку client_mail_vip
    """
    server = smtplib.SMTP("smtp.gmail.com", 587)
    server.starttls()
    file_name = [name_for_save_crop, name_for_save]
    mail_status = ['not_vip', 'vip']
    client_my = []

    try:
        server.login(mail_login, mail_pass)
        msg = MIMEMultipart()
        msg["From"] = mail_login
        msg["Subject"] = "Новый отчет " + str(datetime.today().strftime("%Y-%m-%d") + " для моих подписчиков")
        for file, mail_client_key in zip(file_name, mail_status):
            filename = os.path.basename(file)
            ftype, encoding = mimetypes.guess_type(file)
            filetype, subtype = ftype.split("/")
            with open(f"{file}", 'rb') as f:
                file = MIMEApplication(f.read(), subtype)
            file.add_header('content-disposition', 'attachment', filename=filename)
            msg.attach(file)
            client_my.append(mail_client_key)
            for cli in tqdm(mail_global_dict[mail_client_key]):
                msg['To'] = cli
                server.sendmail(mail_login, cli, msg.as_string())
                time.sleep(1)
                print(f'list [{mail_client_key}] client [{cli}]')
                client_my.append(cli)

        ''' вот тут можно сделать вставку - следующий цикл по отправке почты - сначала по первому списку получателей, 
        а потом суммарное письмо по второму, причем список клиентов можно зипануть в связке с файлом'''
        # for cli in tqdm(client_mail_vip):
        #     msg['To'] = cli
        #     server.sendmail(mail_login, cli, msg.as_string())
        #     time.sleep(2)
        #     client_my.append(cli)
        return f"the message was send -{client_my} "
    except Exception as _ex:
        save_log(prj_path, str(_ex))
        return f"{_ex}\n Check your login"


def dmitry_hist_tab(prj_path, *stope_1):
    tab_name = 'history-all.xlsx'
    dmitry_tab = pd.read_excel(str(prj_path + tab_name), sheet_name='SBP',
                               index_col=0)  ### кстати тут кривое название
    print(dmitry_tab.tiker[pd.DataFrame.notna(dmitry_tab.tiker)])
    dmitry_list_spb = dmitry_tab.tiker[pd.DataFrame.notna(dmitry_tab.tiker)]


def main():
    global prj_path, history_path
    print(f'программа расчета таблицы по стикерам v.{current_version}')
    start_timer = datetime.today()
    print(f'[{start_timer}] START')
    # My constant list
    db_connection_hist_bks = 'mysql+pymysql://python:python@192.168.0.118/history_bks'  # создана база данных history_bks - для сохранения  и работы со сделками,,, надо сделать таблицу, для загрузки итоговой сводной таблицы
    if os.name == 'nt':
        prj_path = path_win
        history_path = path_history_win
    else:
        prj_path = path_linux
        history_path = path_history_linux
    ''' end constant list '''
    save_log(prj_path, '------------ start normal ------------')
    # TEST MODUL

    # кстати выявлено, что в базе данных есть тикеры , есть они в my_st_list, но в итоговой базе они пропадают.
    # надо придумать способ проверять входные и выходные данные, и какой то делать отчет о целостности.
    # может сваять все в одну таблицу - имя, тикер, отрасль, валюта, рынок, последняя дата загрузки, начальная даты загрузки, номер строчки в итоговой таблице в каждой закладке.

    # exit()
    # year_live_stock() # модуля для подсчета годового роста за 12-- месяцев.
    # End TEST MODUL
    # pd_df_to_sql(big_df_table)

    # history_updater(prj_path, sql_login)  # запускаем загрузку и обновление sql базы делает --- отдельный скрипт
    name_for_save = sql_base_make(prj_path, sql_login, col_list)  # делаем расчеты для заполнения таблицы
    name_for_save, name_for_save_crop = excel_format(prj_path, name_for_save, history_path)  # форматируем таблицу

    print(send_email(name_for_save, name_for_save_crop))  # отправляем по почте
    # remove old file after sendmail
    # os.remove(name_for_save)
    # os.remove(name_for_save_crop)
    # save_log(prj_path, str(f'Now remove old file [{os.pardir} {name_for_save}]'))
    # save_log(prj_path, str(f'Now remove old file [{os.pardir} {name_for_save_crop}]'))
    message = f" all Calculating for [{str((datetime.today() - start_timer).seconds)}] sec [{timedelta(seconds=((datetime.today() - start_timer).seconds))}]"
    # ((datetime.today() - start_timer).seconds)}'
    print(datetime.today(), message)
    save_log(prj_path, message)
    exit()

    # дальше пошло старье....

    # exit()
    # stock_name_table(prj_path) # склеиваем все тикеры в одну таблицу  stock_name для базы данных - и кладем в  .csv  .xlsx
    # history_data(prj_path)# загружаем тикеры из таблицы с market_id и загружаем в таблицу stock_hist_data


if __name__ == "__main__":
    main()

    # my mysql request
    ## ALTER IGNORE TABLE hist_data ADD UNIQUE ( Date, st_id(6));  ## удаляем дубликаты в mysql
    # SELECT * FROM hist_data WHERE st_id = 'A';  ## выделяем строки в столбце, где st_id = A
    ## SELECT * FROM hist_data WHERE st_id = 'A' AND open > 70;
    ## SELECT * FROM hist_data WHERE st_id = 'A' AND open > 70 AND date > '2021-08-09';
    # SELECT Date, Low, High, st_id FROM hist_data WHERE st_id = 'APA' AND date > '2021-08-09';
    ##  SELECT Date, Low, High FROM hist_data WHERE st_id = 'APA' AND date BETWEEN '2021-08-09' AND '2021-09-10';
    # """"
    # пробуем получить список тикеров с максимальными датами
    # df = pd.read_sql('SELECT'+ '*' + 'FROM hist_data', con=db_connection)
    #     ##Select st_id, max(date) as date_max from hist_data group by st_id;
    #
    #
    #     delete FROM hist_data WHERE st_id is null;
    #
    #
    # """
    # my mysql request end###

# CREATE TABLE tiker_branch
# (
# ind INT,
# st_id VARCHAR(6) NOT NULL,
# name VARCHAR (70) NOT NULL,
# branch VARCHAR (50) NOT NULL,
# Currency VARCHAR(3) NOT NULL,
# market VARCHAR(3) NOT NULL
# );
