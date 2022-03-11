﻿import mimetypes
import os
import smtplib
import time
from datetime import date
from datetime import datetime
from datetime import timedelta
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from openpyxl.styles import PatternFill
import investpy
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from pandas import DataFrame
from sqlalchemy import create_engine
from tqdm import tqdm

'''
делаем массив для анализа ошибок!!

v22.2 - изменения в связи с заменой источника branch - теперь не из файла, а из базы данных, но там пока нет для US почему то..
!
v22.3 - изменения - оптимизируем код  convert_df_to_np_from_sql - сокращаем число строк и переменные подгружаем в списки
        добавил окраску ячеек в экселе по списку из D_list. 
v22.4 - изменения... добавляем модуль мониторинга.... возможно удалим very_vol




        --- требуется - написать программу, которая будет по списку... (список вручную) .. из базы данных вырезать все значения тикера 
        у которого прошел сплит..  и загружать новые данные за всю историю.
  
'''


def my_hist_data_from_spb_list_update(db_connection_str):
    '''подпрограмма изменения market c US на SPB в базе данных на основании ListingSecurityList.csv
    1) необходимо просто подсунуть ListingSecurityList.csv и все (проверить не поменялось ли название столбца с
    названиями тикеров... дальше все будет все само.
    '''
    # db_connection_str = 'mysql+pymysql://python:python@192.168.0.118/hist_data'
    db_connection = create_engine(db_connection_str)  # connect to database
    df_last_update = pd.read_sql('Select * from base_status ;', con=db_connection)
    df_spb = pd.read_csv('ListingSecurityList.csv', delimiter=';')
    df_spb = df_spb[df_spb['s_currency_kotir'] == 'USD']['s_RTS_code']
    print(f"spb birga {len(df_spb)}")
    spb_df_last = df_last_update[df_last_update['market'] == 'US']['st_id']
    print(f'my df {len(spb_df_last)}')
    listNew = []
    for index in df_spb:
        if spb_df_last.isin([index]).any():
            listNew.append(index)
    print(len(listNew))
    if len(listNew) != 0:
        print(len(listNew))
        print(listNew)
        for index in listNew:
            db_connection.execute(f"update hist_data set market='SPB' where st_id = '{index}';")
            print(index)
        print(f"UPdate Complite. {len(listNew)} st_id market from [US] to [SPB]")


def insert_history_date_into_sql():
    """

    :return:
    """
    df_only_us_last_update = pd.read_sql(
        'Select st_id, max(date) as date_max, Currency, market from hist_data group by st_id',
        con=db_connection)
    market_name = ['United States', 'United States', 'russia']
    stocks_us_investpy = investpy.get_stocks(country=market_name[0])['symbol']
    df_last_update = pd.read_sql('Select * from base_status ;', con=db_connection)
    # df_last_update = pd.read_excel('base_status.xlsx')['st_id']
    df = df_last_update.to_numpy().tolist()
    my_2_list = stocks_us_investpy.to_numpy().tolist()
    my_only_US_df_list = list()
    for index_us in tqdm(my_2_list):
        # print(index_us)
        if index_us in df:
            # print(f'is in -={index_us}')
            continue
        else:
            my_only_US_df_list.append(index_us)
            # print(f"{index_us} - add" )

    print(f"my list len = {len(my_only_US_df_list)}")
    print(f'all US list len = {len(my_2_list)}')
    my_only_US_df_list.sort()

    print(my_only_US_df_list)
    from_date_m, to_date_m = '4/02/2020', '6/02/2020'
    for only_us_index in tqdm(my_only_US_df_list):
        try:
            df_update = investpy.get_stock_historical_data(stock=only_us_index,
                                                           country=market_name[0],
                                                           from_date=from_date_m,
                                                           to_date=to_date_m)
            time.sleep(1.0)
            df_update[['st_id', 'Currency', 'market']] = only_us_index, "USD", 'US'
        except:
            print(f'Error [{only_us_index}] loading')
            continue
        pd_df_to_sql(df_update)


def stock_name_table(linux_path):
    ru_stos1 = investpy.stocks.get_stocks(country='russia')
    ru_stos = ru_stos1[['name', 'symbol']]
    mmm = np.zeros((len(ru_stos), 1)) + 3
    my_id = pd.DataFrame(data=mmm, columns=['markets_id'])
    ru_stos = ru_stos.join(my_id)
    # print(ru_stos)
    big_df = pd.DataFrame(columns=['name', 'symbol'])
    big_df = big_df.append(other=ru_stos)
    df_spb = pd.read_csv('ListingSecurityList.csv', delimiter=';',
                         encoding='cp1251')  # читаем список всех тикеров на СПБ
    stock_spb = df_spb[df_spb['s_currency_kotir'] == 'USD']  # вычисляем список эмитентов СПБ с курсом в USD
    us_st_spb = stock_spb[['name', 'symbol']]  # срезаем лишние столбцы из списка -- надо менять название столбца!!!!!
    mmm = np.zeros((len(df_spb), 1)) + 2
    my_id = pd.DataFrame(data=mmm, columns=['markets_id'])
    us_st_spb = us_st_spb.join(my_id)
    # print(us_st_spb)
    us_stos_usa1 = investpy.stocks.get_stocks(country='United States')
    us_stos_usa = us_stos_usa1[['name', 'symbol']]

    mmm = np.zeros((len(us_stos_usa), 1)) + 1
    my_id = pd.DataFrame(data=mmm, columns=['markets_id'])
    us_stos_usa = us_stos_usa.join(my_id)
    # print(us_stos_usa)
    big_df = big_df.append(other=us_st_spb, ignore_index=True)
    big_df = big_df.append(other=us_stos_usa, ignore_index=True)
    # print('all', big_df)

    with pd.ExcelWriter(linux_path + 'my_all_st.xlsx') as writer:
        big_df.to_excel(writer, sheet_name='all')  ### работает!!!
    # big_df.to_csv('my_test_all_st.csv', sep=';', encoding='cp1251', line_terminator='/n', index=True)

    exit()


def very_vol_stock(df, stok_name, st_id, col_list, branch):  # создаем понедельный список самых волатильных акций
    ''' расчет волатильности - отношения максимумов за период к минимумам,
    !!! данные значения кстати получаются пр ирасчетах основной таблицы - для сокращения времение - можно их взять оттуда и деление произвести на месте'''
    time_list = [-1, -5, -10, -15, -20, -30, -40, -50, -60]
    today_close = df.iloc[-1]['close']
    day_close = df.iloc[-1]['date'].date()
    day_start = df.iloc[0]['date'].date()
    vol = []
    if df.iloc[-10:-1]['close'].min() < 3.0:
        my_round = 6
    else:
        my_round = 2
    for index in range(len(time_list) - 1):
        vol.append( round( ( df.iloc[time_list[index + 1]:time_list[index]]['high'].max()/df.iloc[time_list[index + 1]:time_list[index]]['low'].min() - 1) * 100, my_round) )
    my_frame = [st_id, stok_name, branch, today_close, *vol,
                day_start, day_close]
    return pd.DataFrame([my_frame], columns=list(col_list))

def c_200_df_make(df, stiker, name, c_200_col,
                  branch):  # Модуль подсчета скользящей средней 200 дней на периоды до 85 дней назад
    df.set_index('date', inplace=True)
    np_array = df[['high', 'low', 'close']].to_numpy()
    today_close = np_array[-1, 2]
    day_close = df.index[-1].date()
    day_start = df.index[0].date()
    if np.mean(np_array[-10: -1, 2]) < 3.0:
        my_round = 6
    else:
        my_round = 2
    c_200_0 = np.mean(np_array[-200:-1, 0:2]).round(my_round)
    c_200_1 = np.mean(np_array[-205:-5, 0:2]).round(my_round)
    c_200_2 = np.mean(np_array[-215:-15, 0:2]).round(my_round)
    c_200_3 = np.mean(np_array[-225:-25, 0:2]).round(my_round)
    c_200_4 = np.mean(np_array[-235:-35, 0:2]).round(my_round)
    c_200_5 = np.mean(np_array[-245:-45, 0:2]).round(my_round)
    c_200_6 = np.mean(np_array[-255:-55, 0:2]).round(my_round)
    c_200_7 = np.mean(np_array[-265:-65, 0:2]).round(my_round)
    c_200_8 = np.mean(np_array[-285:-110, 0:2]).round(my_round)

    pr_mean_0 = np.mean(np_array[-1, 0:2])
    pr_mean_1 = np.mean(np_array[-5, 0:2])
    pr_mean_2 = np.mean(np_array[-15, 0:2])
    pr_mean_3 = np.mean(np_array[-25, 0:2])
    pr_mean_4 = np.mean(np_array[-35, 0:2])
    pr_mean_5 = np.mean(np_array[-45, 0:2])
    pr_mean_6 = np.mean(np_array[-55, 0:2])
    pr_mean_7 = np.mean(np_array[-65, 0:2])
    pr_mean_8 = np.mean(np_array[-110, 0:2])
    delta0_0 = pr_mean_0 - c_200_0
    delta1_1 = pr_mean_1 - c_200_1
    delta2_2 = pr_mean_2 - c_200_2
    delta3_3 = pr_mean_3 - c_200_3
    delta4_4 = pr_mean_4 - c_200_4
    delta5_5 = pr_mean_5 - c_200_5
    delta6_6 = pr_mean_6 - c_200_6
    delta7_7 = pr_mean_7 - c_200_7
    delta8_8 = pr_mean_8 - c_200_8
    del_pr0 = round(100 * delta0_0 / pr_mean_0, 2)
    del_pr1 = round(100 * delta1_1 / pr_mean_1, 2)
    del_pr2 = round(100 * delta2_2 / pr_mean_2, 2)
    del_pr3 = round(100 * delta3_3 / pr_mean_3, 2)
    del_pr4 = round(100 * delta4_4 / pr_mean_4, 2)
    del_pr5 = round(100 * delta5_5 / pr_mean_5, 2)
    del_pr6 = round(100 * delta6_6 / pr_mean_6, 2)
    del_pr7 = round(100 * delta7_7 / pr_mean_7, 2)
    del_pr8 = round(100 * delta8_8 / pr_mean_8, 2)

    my_frame = [stiker, name, branch, today_close, c_200_0, c_200_1, c_200_2, c_200_3, c_200_4, c_200_5, c_200_6,
                c_200_7, c_200_8, pr_mean_0, pr_mean_1, pr_mean_2, pr_mean_3, pr_mean_4, pr_mean_5, pr_mean_6,
                pr_mean_7, pr_mean_8, delta0_0, delta1_1, delta2_2, delta3_3, delta4_4, delta5_5, delta6_6,
                delta7_7, delta8_8, day_start, day_close, del_pr0, del_pr1, del_pr2, del_pr3, del_pr4, del_pr5, del_pr6,
                del_pr7, del_pr8]
    return pd.DataFrame([my_frame], columns=list(c_200_col))

def year_live_stock():  # df, stick_id, st_name): #
    ''' расчет годового роста каждый месяц -- еще не доделано!!!!! '''
    print('Блок расчета годоводго роста по месяцам')
    df_col_list = ['name', 'stiker', 'today_close', 'inter1', 'inter2', 'inter3', 'inter4', 'inter5', 'inter6',
                   'inter7', 'inter8', 'inter9', 'inter10', 'inter11', 'inter12', 'inter_sum', 'inter_mean']
    '''' inter1 - интервал с даты отчета - на год назад, inter12 -- интервал с года назад до 2 лет назад, остальные промежуточные 
    за базу расчетов берем среднее из максимумов и минимумов за десять дней до даты отсчета'''
    df_date = pd.DataFrame(
        columns=['date_start_now', 'date_end_now', 'date_start_12', 'date_end_12'])  # создали датафрейм со всеми датами
    stick_id, st_name = 'AAPL', 'Apple'
    db_connection_str = 'mysql+pymysql://python:python@192.168.0.118/hist_data'
    df_from_sql = pd.read_sql(
        'Select date, high, low, close, st_id, Currency from hist_data  WHERE Currency=\'USD\' and st_id = \'AAPL\';',
        con=db_connection_str)  ## загружаем базу СПБ ---USD
    k, y_q, m_d, m_m = 0, 0, 0, 0
    cur_date = datetime.today()
    for i in range(0, 12):
        if date.today().month - i + k <= 0:
            k, y_q = 12, 1
        if date.today().day - 10 < 0:
            m_d, m_m = 29, -1
        df_date.loc[i, 'date_start_now'] = date(date.today().year - y_q, date.today().month - i + k + m_m,
                                                date.today().day - 10 + m_d)
        df_date.loc[i, 'date_end_now'] = date(date.today().year - y_q, date.today().month - i + k, date.today().day)
        df_date.loc[i, 'date_start_12'] = date(date.today().year - y_q - 1, date.today().month - i + k + m_m,
                                               date.today().day - 10 + m_d)
        df_date.loc[i, 'date_end_12'] = date(date.today().year - y_q - 1, date.today().month - i + k, date.today().day)
    # print(df_date.tail(3))
    # print(df_from_sql.tail(10))
    delta_np = np.zeros(12)
    print(delta_np)
    indexx = 0
    for indexx in range(len(df_date)):
        # print(df_date.date_start_now[indexx],df_date.date_end_now [indexx] )
        mean_delta = 0.5 * (pd.DataFrame.mean(
            df_from_sql[df_from_sql.date > pd.to_datetime(df_date.date_start_now[indexx], format='%Y-%m-%d')].iloc[:10][
                'high'])
                            + pd.DataFrame.mean(df_from_sql[
                                                    df_from_sql.date > pd.to_datetime(df_date.date_start_now[indexx],
                                                                                      format='%Y-%m-%d')].iloc[:10][
                                                    'low']))
        mean_delta2 = 0.5 * (pd.DataFrame.mean(
            df_from_sql[df_from_sql.date > pd.to_datetime(df_date.date_start_12[indexx], format='%Y-%m-%d')].iloc[:10][
                'high'])
                             + pd.DataFrame.mean(
                    df_from_sql[
                        df_from_sql.date > pd.to_datetime(df_date.date_start_12[indexx], format='%Y-%m-%d')].iloc[
                    :10]['low']))
        delta_np[indexx] = round((mean_delta - mean_delta2) / mean_delta2, 3)
        # print('d1',round(mean_delta,2),'d2', round(mean_delta2,2), 'delta_pr', delta_np[indexx])
    print('delta', delta_np)
    print('mean', round(delta_np.mean(), 3))
    print('sum', delta_np.sum())

    year_live_stock_df = pd.DataFrame(columns=df_col_list)
    #    year_live_stock_df = year_live_stock_df[]

    exit()

def save_history(linux_path, big_df_table):
    big_df_table.to_excel(linux_path + 'history_data.xlsx')

def save_log(linux_path, message):  # сохраняет в лог файл сообщение..
    f = open(linux_path + 'make_from_sql.log', mode='a')
    lines = '[' + str(datetime.today()) + '] ' + str(message + ' v.22.4')
    f.writelines(lines + '\n')
    f.close()
    # print(lines+ '\n')

def history_updater(linux_path, db_connection_str):  ## модуль делает обновление исторических данных тикеров базы mysql

    cur_date = datetime.today()
    print('сегодня...', cur_date.strftime("%Y-%m-%d"), 'Now start UPDATE mysql')
    market_name = ['United States', 'United States', 'russia']

    db_connection = create_engine(db_connection_str)  # connect to database
    df_last_update = pd.read_sql('Select * from base_status ;', con=db_connection)
    print("load [base status] from Mysql...[OK]")

    # print('длина массива',len(df_last_update))
    for ind in tqdm(range(len(df_last_update))):
        deltadays = (cur_date - df_last_update.iloc[ind, 1]).days
        if deltadays <= 20 and deltadays > 1:
            # print(':',df_last_update.iloc[ind,0],':', 'last date-', (df_last_update.iloc[ind,1]).strftime("%Y-%m-%d"), ',дней без обновления-', deltadays )
            # print('from_date:', (df_last_update.iloc[ind,1]- timedelta(days=1) ).strftime("%Y-%m-%d"), ',to_Date:',cur_date.strftime("%Y-%m-%d") )
            if df_last_update.iloc[ind, 2] == 'USD':
                try:
                    df_update = investpy.get_stock_historical_data(df_last_update.iloc[ind, 0], country=market_name[0],
                                                                   from_date=(df_last_update.iloc[ind, 1] - timedelta(
                                                                       days=1)).strftime("%d/%m/%Y"),
                                                                   to_date=cur_date.strftime("%d/%m/%Y"))
                    # print("market=", market_name[0])
                    mmm = np.ones((len(df_update), 1))  # len_mmm
                    st_id = pd.DataFrame(data=mmm, columns=['st_id'])
                    st_id['st_id'] = df_last_update.iloc[ind, 0]
                    df_update.reset_index(level=['Date'], inplace=True)  # замена индекса
                    df_update = df_update.join(pd.DataFrame(st_id, columns=['st_id']))
                    # print('pppll',df_last_update.iloc[ind,0],df_update)
                    pd_df_to_sql(df_update, db_connection_str, df_last_update.iloc[ind, 0], ind, len(df_last_update))
                except:
                    print('[', datetime.today(), ']', "investpy load error[", df_last_update.iloc[ind, 0], ']')
                    continue
            elif df_last_update.iloc[ind, 2] == 'RUB':
                try:
                    df_update = investpy.get_stock_historical_data(df_last_update.iloc[ind, 0], country=market_name[2],
                                                                   from_date=(timedelta(days=1) + df_last_update.iloc[
                                                                       ind, 1]).strftime("%d/%m/%Y"),
                                                                   to_date=cur_date.strftime("%d/%m/%Y"))
                    # print("market=", market_name[2])
                    mmm = np.ones((len(df_update), 1))  # len_mmm
                    st_id = pd.DataFrame(data=mmm, columns=['st_id'])
                    st_id['st_id'] = df_last_update.iloc[ind, 0]
                    df_update.reset_index(level=['Date'], inplace=True)  # замена индекса
                    df_update = df_update.join(pd.DataFrame(st_id, columns=['st_id']))
                    # print(df_update)
                    pd_df_to_sql(df_update, db_connection_str, df_last_update.iloc[ind, 0], ind, len(df_last_update))
                except:
                    print('[', datetime.today(), ']', "investpy load error[", df_last_update.iloc[ind, 0], ']')
                    continue
        time.sleep(1)
    print("Mysql base is .... [UPDATE]")

def input_tables_append(db_connection_str, linux_path):
    st_name = pd.read_excel(linux_path + 'my_all_st.xlsx', index_col=0)
    branch_name = pd.read_excel(linux_path + 'tiker-branch.xlsx', index_col=0)
    # db_connection = create_engine(db_connection_str)
    # df_last_update = pd.read_sql('Select st_id, max(date) as date_max, Currency  from hist_data group by st_id',                                 con=db_connection)
    col_list_m = ['name', 'tiker', 'branch', 'currency', 'market', 'max_date', 'min_date', 'status_table']
    print(st_name.head(3), len(st_name))

    print(branch_name.head(3), len(branch_name))
    # print(df_last_update.head(3), len(df_last_update))
    status_t = pd.DataFrame(columns=col_list_m)
    print(status_t)
    status_index = 0
    for ind in range(len(branch_name.st_id)):
        print(branch_name['st_id'].iloc[ind])
        if pd.DataFrame.any(status_t['st_id'] != branch_name['st_id'].iloc[ind]):
            status_t['st_id'].iat[status_index] = branch_name['st_id'].iloc[ind]
            status_index += 1

    print(status_t, status_index)

def convert_df_to_np_from_sql(df1, stok_name, st_inv, col_list, branch,
                              df_teh):  # подчтет данных их исторических значений
    ''' оптимизировал код - исключен нумпай и прочие строчки'''
    time_list = [-1, -5, -10, -15, -20, -30, -40, -50, -60]
    today_close = df1.iloc[-1]['close']
    day_close = df1.iloc[-1]['date'].date()
    day_start = df1.iloc[0]['date'].date()
    # df1.index[0].date()
    min_y = df1.iloc[:]['low'].min()
    max_y = df1.iloc[:]['high'].max()
    min_y_date = df1[df1['low'] == min_y].iloc[0]['date'].date()
    max_y_date = df1[df1['high'] == max_y].iloc[0]['date'].date()
    if df1.iloc[-10:-1]['close'].min() < 3.0:
        my_round = 6
    else:
        my_round = 2
    min_dek, max_dek = [], []
    min_per, max_per = [], []
    min_pr_delta, max_pr_delta = [], []
    for index in range(len(time_list) - 1):
        min_dek.append((df1.iloc[time_list[index + 1]:time_list[index]]['low'].min()).round(my_round))
        max_dek.append((df1.iloc[time_list[index + 1]:time_list[index]]['high'].max()).round(my_round))
    for index_2 in range(0, 8):
        min_per.append(round((today_close / min_dek[index_2] - 1) * 100, my_round))
        max_per.append(round((today_close / max_dek[index_2] - 1) * 100, my_round))
    for index in range(0, 7):
        min_pr_delta.append(round(min_per[index] - min_per[index + 1], my_round))
        max_pr_delta.append(round(max_per[index] - max_per[index + 1], my_round))
        # print( df1.iloc[time_list[index + 1]:time_list[index]]['low'].min(), index, time_list[index], time_list[index + 1])
    today_y_pr_max = round((today_close / max_y - 1) * 100, 1)
    today_y_pr_min = round((today_close / min_y - 1) * 100, 1)
    m1_max, m1_min = df1.iloc[-20:-1]['high'].max(), df1.iloc[-20:-1]['low'].min()
    m3_max, m3_min = df1.iloc[-60:-1]['high'].max(), df1.iloc[-60:-1]['low'].min()
    m6_max, m6_min = df1.iloc[-120:-1]['high'].max(), df1.iloc[-120:-1]['low'].min()
    year1_max, year1_min = df1.iloc[-240:-1]['high'].max(), df1.iloc[-240:-1]['low'].min()
    pr_30_day_max, pr_30_day_min = round((today_close / m1_max - 1) * 100, 1), round((today_close / m1_min - 1) * 100,
                                                                                     1)
    pr_90_day_max, pr_90_day_min = round((today_close / m3_max - 1) * 100, 1), round((today_close / m3_min - 1) * 100,
                                                                                     1)
    pr_6_m_max, pr_6_m_min = round((today_close / m6_max - 1) * 100, 1), round((today_close / m6_min - 1) * 100, 1)
    pr_1y_max, pr_1y_min = round((today_close / year1_max - 1) * 100, 1), round((today_close / year1_min - 1) * 100, 1)
    my_frame = [st_inv, stok_name, branch, today_close, *min_dek, *min_per, *min_pr_delta,
                *max_dek, *max_per, *max_pr_delta,
                min_y, max_y, today_y_pr_min, today_y_pr_max, day_start, day_close, *df_teh,
                m1_max, m1_min, m3_max, m3_min, m6_max, m6_min, year1_max, year1_min, pr_30_day_max, pr_30_day_min,
                pr_90_day_max, pr_90_day_min,
                pr_6_m_max, pr_6_m_min, pr_1y_max, pr_1y_min]
    # , min_date, max_date]
    df2 = pd.DataFrame([my_frame], columns=list(col_list))
    return df2  # готовим данные для записи в таблицу


def sql_base_make(linux_path, db_connection_str,
                  col_list):  # Модуль загрузки данных из базы mysql и формирования отчетных таблиц
    global branch_name_local
    db_connection = create_engine(db_connection_str)  # connect to database
    big_df: DataFrame = pd.DataFrame(columns=list(col_list))
    big_df_US: DataFrame = pd.DataFrame(columns=list(col_list))
    today_date = datetime.today()
    very_vol_list = ['stiker', 'name', 'branch', 'today_close', 'vol_1', 'vol_2', 'vol_3', 'vol_4', 'vol_5', 'vol_6',
                     'vol_7', 'vol_8', 'day_start', 'day_close']  # список колонок для списка волатильных акций
    c_200_col = ['stiker', 'name', 'branch', 'today_close', 'c_200_0', 'c_200_1', 'c_200_2', 'c_200_3', 'c_200_4',
                 'c_200_5', 'c_200_6', 'c_200_7', 'c_200_8', 'pr_mean_0', 'pr_mean_1', 'pr_mean_2', 'pr_mean_3',
                 'pr_mean_4', 'pr_mean_5', 'pr_mean_6',
                 'pr_mean_7', 'pr_mean_8', 'delta0_0', 'delta1_1', 'delta2_2', 'delta3_3', 'delta4_4', 'delta5_5',
                 'delta6_6',
                 'delta7_7', 'delta8_8', 'day_start', 'day_close',
                 'del_pr0', 'del_pr1', 'del_pr2', 'del_pr3', 'del_pr4', 'del_pr5', 'del_pr6', 'del_pr7', 'del_pr8']
    teh_an_list = ['teh_daily_sel', 'teh_daily_buy', 'teh_weekly_sel', 'teh_weekly_buy', 'teh_monthly_sell',
                   'teh_monthly_buy', 'daily_sma_signal 200', 'daily_ema_signal 200', 'weekly_sma_signal 200',
                   'weekly_ema_signal 200', 'monthly_sma_signal 200', 'monthly_ema_signal 200', 'EPS', 'P_E']
    teh_an_df_nodata = pd.DataFrame(columns=teh_an_list)
    teh_an_df_nodata.loc[len(teh_an_df_nodata)] = 'No DAta'
    c_200_df, c_200_df_ru = pd.DataFrame(columns=list(c_200_col)), pd.DataFrame(columns=list(c_200_col))
    very_vol_spb, very_vol_ru = pd.DataFrame(columns=list(very_vol_list)), pd.DataFrame(columns=list(very_vol_list))
    big_df_ru = pd.DataFrame(columns=list(col_list))
    # отрасли
    branch_name = pd.read_sql('Select * from tiker_branch ;', con=db_connection)
    # tiker , name, branch,  curency
    ###
    tab_name = 'history-all.xlsx'  # создаем список собственных тикеров
    client_tiker_list_tab = pd.read_excel(str(linux_path + tab_name), sheet_name='SPB',
                                          index_col=0)  ### кстати тут кривое название
    dmitry_list_spb = set(client_tiker_list_tab.tiker[pd.DataFrame.notna(client_tiker_list_tab.tiker)])
    zina_list_spb = {'AA', 'APA', 'BA', 'FSLR', 'INTC', 'NEE', 'PBF', 'SPR', 'LTHM', 'SPCE', 'UAL', 'NVTA', 'VIR',
                     'NVDA', 'NDAQ', 'U'}
    print(f'D dataframe list  created [{len(dmitry_list_spb)}]')
    print(f'Z dataframe list  created [{len(zina_list_spb)}]')
    print('[', datetime.today(), ']', "Dataframe load...[OK]")
    save_log(linux_path, "Dataframe load...[OK]")

    df_last_update = pd.read_sql('Select * from base_status ;', con=db_connection)
    last_day_sql = df_last_update.iloc[1]['date_max'].date()
    if (today_date.date() - last_day_sql).days != 0:
        df_last_update = pd.read_sql(
            'Select st_id, max(date) as date_max, Currency, min(date) as date_min , market from hist_data group by st_id',
            con=db_connection)  # загрузили список тикеров из базы с последней датой
        df_last_update['today_day'] = datetime.today()
        df_last_update.to_sql(name='base_status', con=db_connection, if_exists='replace')  # append , replace
        save_log(linux_path, 'base_status table is Update ')
    df_last_teh = pd.read_sql('Select st_id, max(date) as date_max from teh_an group by st_id',
                              con=db_connection)
    max_teh_date = pd.DataFrame.max(df_last_teh.date_max[:])
    teh_full = round(len(df_last_teh[df_last_teh.date_max == max_teh_date]) / len(df_last_teh), 2)
    min_teh_date = pd.DataFrame.min(df_last_teh.date_max[:])
    print('teh_Full', teh_full * 100, '% Have max date')
    save_log(linux_path, 'teh_full ' + str(teh_full * 100) + '% have MAX date' )
    if teh_full != 1:
        save_log(linux_path, 'base teh analis not full - only ' + str(teh_full * 100) + ' %')
        max_teh_date = min_teh_date # берем для использования минимальную из максимальных
    save_log(linux_path, 'teh_date - [' + str(max_teh_date) + ']')
    # проходимся по базе - если она старше 10 дней - просим обновиться, заодно удаляем старые тикеры, а также все, что на бирже не обновляется давно
    max_old_days = 10
    max_stick, start_stick_num = len(df_last_update['st_id']), 0
    max_date = pd.DataFrame.max(df_last_update.date_max[:])
    print('max date', max_date)
    for indexx in df_last_update['st_id']:
        if (today_date - df_last_update[df_last_update.st_id == indexx].iloc[0]['date_max']).days <= max_old_days:
            start_stick_num += 1
            # print(indexx)
    message = f'hist_data тикеров моложе [{max_old_days}] дней [{start_stick_num}], всего тикеров [{max_stick}], доля=[{round(start_stick_num / max_stick, 2)}] %'
    save_log(linux_path, message)
    print(message)
    df_out_date = df_last_update[df_last_update.date_max <= max_date - timedelta(days=max_old_days)]
    df_last_update = df_last_update[df_last_update.date_max > max_date - timedelta(
        days=max_old_days)]  ### отрезаем все тикеры , для которых данные старые!!!!!!!!! --- так можно все порезать так, что считать будет нечего
    df_out_date.sort_values(by=['date_max'], inplace=True)
    save_log(linux_path, f' len of outdate of history_date{len(df_out_date)}' )
    df_teh = pd.read_sql(
        'Select hd.* from teh_an hd join (Select hd.st_id, max(hd.date) as date_max from teh_an hd group by hd.st_id) hist_data_date_max on hist_data_date_max.st_id = hd.st_id and hist_data_date_max.date_max = hd.date;',
        con=db_connection)
    my_start_date = '2020-02-01'
    print('[', datetime.today(), ']', "load LIST from Mysql...[OK]")
    print('teh_an shape', df_teh.shape)
    df_spb = pd.read_sql(
        f'Select date, high, low, close, st_id, Currency from hist_data  WHERE market=\'SPB\' and date > \'{my_start_date}\';',
        con=db_connection)  ## загружаем базу СПБ ---USD
    print('SPB sql load OK')
    df_ru = pd.read_sql(
        f'Select date, high, low, close, st_id, Currency from hist_data  WHERE market=\'RU\' and date > \'{my_start_date}\';',
        con=db_connection)  ## загружаем базу СПБ ---RUB
    print('RU sql load OK')
    df_from_us = pd.read_sql(
        f'Select date, high, low, close, st_id, Currency from hist_data  WHERE market=\'US\' and date > \'{my_start_date}\';',
        con=db_connection)  ## загружаем базу US
    print('US sql load OK')
    # считаем актуальность базы данных
    for market_s in df_last_update['market'].unique():
        listing_ll = pd.Series(
            {c: df_last_update[df_last_update['market'] == market_s][c].unique() for c in df_last_update})
        message = f'sorted Listing_ll {listing_ll[1].sort()}'
        print(message); save_log(linux_path,message)
        for num_1 in range(len(listing_ll[2]) - 2, len(listing_ll[2])):
            print(
                f"date for [{market_s}]- [{str(listing_ll['date_max'][num_1])[:10]}] is [{len(df_last_update[(df_last_update['market'] == market_s) & (df_last_update['date_max'] == listing_ll['date_max'][num_1])]['date_max'])}] ")
            save_log(linux_path,
                     f"date for [{market_s}]=[{str(listing_ll['date_max'][num_1])[:10]}] is [{len(df_last_update[(df_last_update['market'] == 'SPB') & (df_last_update['date_max'] == listing_ll['date_max'][num_1])]['date_max'])}] ")
    save_log(linux_path, "SPB base load--" + str(len(df_spb)) + "...[OK]")
    save_log(linux_path, "RU base load--" + str(len(df_ru)) + "...[OK]")
    save_log(linux_path, "US base load--" + str(len(df_from_us)) + "...[OK]")
    print('[', datetime.today(), ']', "load DATA from Mysql...[OK]")
    # print(f'SPB base [{len(df_spb[])}') доделать надо..
    save_log(linux_path, 'load DATA from Mysql...[OK]')
    '''
    # загружаем базу данных, и список тикеров с последней датой обновления . сортируем по дате. по списку
    # тикеров проходимся в базе данных -
    # делаем выборку по тикеру - столбцы date, high, low -- переводим индекс date. конвертируем в numpy.
    # передаем в функцию
    '''
    df_spb.sort_values(by=['date'], inplace=True)
    df_ru.sort_values(by=['date'], inplace=True)
    df_from_us.sort_values(by=['date'], inplace=True)
    start_timer = datetime.today()
    print('\n[', datetime.today(), ']', 'stage 1 (vol).....[Calculating]') # считаем волатильные акции
    for ind in tqdm(df_last_update[df_last_update['st_id']== 'AAPL'] ['st_id']):
        if (branch_name['st_id'] == ind).any():
            branch_name_local = branch_name[branch_name.st_id == ind].iloc[0]['branch']
        else:
            branch_name_local = 'NO data'
        if df_last_update[df_last_update.st_id == ind].iloc[0]['market'] == "SPB":
            df_1 = df_spb[df_spb['st_id'] == ind]
            try:
                df_v_spb = very_vol_stock(df_1, branch_name[branch_name.st_id == ind].iloc[0]['name'], ind,
                                          very_vol_list, branch_name_local)
            except:
                message = 'USD_very_vol error ' + str(ind)
                # save_log( linux_path, message)
                continue
            very_vol_spb = pd.concat([very_vol_spb,df_v_spb])
        elif df_last_update[df_last_update.st_id == ind].iloc[0]['market'] == "RU":
            df_2 = df_ru[df_ru['st_id'] == ind]
            try:
                df_v_ru = very_vol_stock(df_2, branch_name[branch_name.st_id == ind].iloc[0]['name'], ind,
                                         very_vol_list, branch_name_local)
            except:
                message = 'RU_very_vol error ' + str(ind)
                # save_log(linux_path, message)
                continue
            very_vol_ru = pd.concat([very_vol_ru,df_v_ru])
    print('vol-ru', len(very_vol_ru), 'vol- spb', len(very_vol_spb))
    print('\n[', datetime.today(), ']', 'stage 2 (c_200)...[Calculating]')
    for ind in tqdm(df_last_update[df_last_update['st_id']== 'AAPL'] ['st_id']):  # считаем скользящую среднюю
        if pd.DataFrame.any(branch_name.st_id == ind):
            branch_name_local = branch_name[branch_name.st_id == ind].iloc[0]['branch']
        else:
            branch_name_local = 'NO data'
        if df_last_update[df_last_update.st_id == ind].iloc[0]['market'] == "SPB":
            df_1 = df_spb[df_spb['st_id'] == ind]
            try:
                c_200 = c_200_df_make(df_1, ind, branch_name[branch_name.st_id == ind].iloc[0]['name'], c_200_col,
                                      branch_name[branch_name.st_id == ind].iloc[0]['branch'])
            except:
                message = 'USD_c_200 error ' + str(ind)
                # save_log(linux_path, message)
                continue
            c_200_df = pd.concat([c_200_df,c_200]) # , list(c_200_col))
        elif df_last_update[df_last_update.st_id == ind].iloc[0]['market'] == "RU":
            df_2 = df_ru[df_ru['st_id'] == ind]
            try:
                c_200 = c_200_df_make(df_2, ind, branch_name[branch_name.st_id == ind].iloc[0]['name'], c_200_col,
                                      branch_name_local)
            except:
                message = 'RU_c_200 error ' + str(ind)
                # save_log(linux_path, message)
                continue
            c_200_df_ru = pd.concat([c_200_df_ru,c_200]) #, list(c_200_col))
    print('200-ru', len(c_200_df_ru), '200-spb', len(c_200_df))
    print('\n[', datetime.today(), ']', 'stage 3 (SPB).....[Calculating]')
    for ind in tqdm(df_last_update['st_id']):  # считаем рынок СПБ
        if pd.DataFrame.any(branch_name.st_id == ind):
            branch_name_local = branch_name[branch_name.st_id == ind].iloc[0]['branch']
        else:
            branch_name_local = 'NO data'
        if df_last_update[df_last_update.st_id == ind].iloc[0]['market'] == "SPB":
            df_1 = df_spb[df_spb['st_id'] == ind]
            # print('name', st_name[st_name.symbol== ind].iloc[0]['name'])
            try:
                my_df = convert_df_to_np_from_sql(df_1, branch_name[branch_name.st_id == ind].iloc[0]['name'], ind,
                                                  col_list,
                                                  branch_name_local,
                                                  df_teh[df_teh.st_id == ind].iloc[0][3:])  # пробуем добавить теханализ
            except:
                message = 'SPB error ' + str(ind)
                save_log(linux_path, message)
                continue
            big_df = pd.concat([big_df,my_df])
    print('\n[', datetime.today(), ']', 'stage 4 (RU)......[Calculating]')
    for ind in tqdm(df_last_update['st_id']):  # считаем рынок Московская биржа
        if pd.DataFrame.any(branch_name.st_id == ind):
            branch_name_local = branch_name[branch_name.st_id == ind].iloc[0]['branch']
        else:
            branch_name_local = 'NO data'
        if df_last_update[df_last_update.st_id == ind].iloc[0]['market'] == "RU":
            df_2 = df_ru[df_ru['st_id'] == ind]
            try:
                my_df_ru = convert_df_to_np_from_sql(df_2, branch_name[branch_name.st_id == ind].iloc[0]['name'],
                                                     ind, col_list, branch_name_local,
                                                     df_teh[df_teh.st_id == ind].iloc[0][3:], )
            except:
                message = 'RU error ' + str(ind)
                save_log(linux_path, message)
                continue
            big_df_ru = pd.concat([big_df_ru,my_df_ru])
    print('\n[', datetime.today(), ']', 'stage 5 (dmitry)..[Collecting]')
    ###  создаем сборку для фильтрации – это простая выборка из массива на основе шаблона – гораздо проще чем поиск
    # кстати надо попробовать предыдущие циклы переделать с помощью команды map – наверное не получится – слишком много переменных передавать надо
    ### зато интересный скилл изучили..!!!
    dmitry_list_for_filter = big_df['tiker'].isin([*dmitry_list_spb])
    dmitry_df = big_df[dmitry_list_for_filter].copy()
    print('\n[', datetime.today(), ']', 'stage 6 (zina)..[Collecting]')
    zina_df = big_df[big_df['tiker'].isin ([*zina_list_spb])].copy()
    if datetime.today().weekday() == 5:
        print('\n[', datetime.today(), ']', 'stage 7 (US)..[Collecting]')
        for tik_index in tqdm(df_last_update['st_id']):
            if pd.DataFrame.any(branch_name.st_id == tik_index):
                branch_name_local = branch_name[branch_name.st_id == tik_index].iloc[0]['branch']
            else:
                branch_name_local = 'NO data'
            if df_last_update[df_last_update.st_id == tik_index].iloc[0]['market'] == "US":
                df_1 = df_from_us[df_from_us['st_id'] == tik_index]
                # print('name', st_name[st_name.symbol== ind].iloc[0]['name'])
                try:
                    my_df = convert_df_to_np_from_sql(df_1, branch_name[branch_name.st_id == tik_index].iloc[0]['name'], tik_index,
                                                      col_list,
                                                      branch_name_local,
                                                      df_teh[df_teh.st_id == tik_index].iloc[0][3:])  # пробуем добавить теханализ
                except:
                    message = 'US error ' + str(tik_index)
                    save_log(linux_path, message)
                    continue
                big_df_US = big_df_US.append(my_df)  # , list(col_list))
    else:
        print('\n[', datetime.today(), ']', 'today not for stage 7 -- (US), see later -(need 5 day of weekday)')

    name_for_save = str(linux_path) + 'sql_make-' + str(start_timer.date()) + '.xlsx'
    with pd.ExcelWriter(name_for_save) as writer:  # записываем отчетный файл
        big_df.to_excel(writer, sheet_name='SPB')  # считаем весь рынок SPB
        big_df_ru.to_excel(writer, sheet_name='RU')  # считаем весь рынок RU
        big_df_US.to_excel(writer, sheet_name='USA')  # считаем весь рынок USA
        c_200_df.to_excel(writer, sheet_name='C200_SPB')  # считаем скользящие средние
        c_200_df_ru.to_excel(writer, sheet_name='C200_RU')  # считаем скользящие средние
        very_vol_spb.to_excel(writer, sheet_name='very_vol_SPB')  # считаем волатильность
        very_vol_ru.to_excel(writer, sheet_name='very_vol_RU')  # считаем волатильность
        df_teh.to_excel(writer, sheet_name='Teh_analis_all')  #
        dmitry_df.to_excel(writer, sheet_name='D_list')
        zina_df.to_excel(writer, sheet_name='Z_list')
        # считаем весь рынок + волатильность RU
        # считаем весь рынок + волатильность SPB

    print('\n[', datetime.today(), ']', 'EXCEL file .... [SAVED]')
    print('\n[', datetime.today(), ']', (datetime.today() - start_timer).seconds, '[sec]')
    save_log(linux_path, '\n[' + str(datetime.today()) + ']' + str((datetime.today() - start_timer).seconds) + '[sec]')
    return name_for_save

def my_start():  # исходные данные - константы
    col_list = ['tiker', 'name', 'branch', 'today_close',
                'mean_min_dek1', 'mean_min_dek2', 'mean_min_dek3', 'mean_min_dek4',
                'mean_min_dek5', 'mean_min_dek6', 'mean_min_dek7', 'mean_min_dek8',
                'mean_min_pr_dek1',
                'mean_min_pr_dek2', 'mean_min_pr_dek3', 'mean_min_pr_dek4', 'mean_min_pr_dek5', 'mean_min_pr_dek6',
                'mean_min_pr_dek7', 'mean_min_pr_dek8',
                'min_pr_delta_1_2', 'min_pr_delta_2_3', 'min_pr_delta_3_4', 'min_pr_delta_4_5', 'min_pr_delta_5_6',
                'min_pr_delta_6_7', 'min_pr_delta_7_8',
                'max_dek1', 'max_dek2', 'max_dek3', 'max_dek4', 'max_dek5',
                'max_dek6', 'max_dek7', 'max_dek8',
                'max_pr_dek1', 'max_pr_dek2', 'max_pr_dek3', 'max_pr_dek4',
                'max_pr_dek5', 'max_pr_dek6', 'max_pr_dek7', 'max_pr_dek8',
                'max_pr_delta_1_2', 'max_pr_delta_2_3', 'max_pr_delta_3_4', 'max_pr_delta_4_5', 'max_pr_delta_5_6',
                'max_pr_delta_6_7', 'max_pr_delta_7_8',
                'min_y', 'max_y', 'today_y_pr_min',
                'today_y_pr_max', 'day_start', 'day_close',
                # """,
                'teh_daily_sel', 'teh_daily_buy', 'teh_weekly_sel', 'teh_weekly_buy', 'teh_monthly_sell',
                'teh_monthly_buy', 'daily_sma_signal 200', 'daily_ema_signal 200', 'weekly_sma_signal 200',
                'weekly_ema_signal 200', 'monthly_sma_signal 200', 'monthly_ema_signal 200', 'EPS', 'P_E',
                'm1_max', 'm1_min', 'm3_max', 'm3_min', ' m6_max', 'm6_min', 'year1_max', 'year1_min', 'pr_30_day_max',
                'pr_30_day_min',
                'pr_90_day_max', 'pr_90_day_min',
                'pr_6_m_max', 'pr_6_m_min', 'pr_1y_max', 'pr_1y_min'

                ]
    # """
    return col_list

def excel_format(linux_path, name_for_save, history_path):  # форматирование отчетного файла после записи
    redFill = PatternFill(start_color='FFFF0000',
                          end_color='FFFF0000',
                          fill_type='solid')

    wb_help = load_workbook(filename=str(linux_path) + 'df_help_06.03.2022.xlsx')  # шаблон форматирования
    shit_n = wb_help.sheetnames
    ws = wb_help.active
    print('[', datetime.today(), ']', 'EXCEL format [START]')
    print(shit_n)
    # for s_name in tqdm(shit_n):
    #     sheet = wb_help[s_name]
    #     m_row = sheet.max_row
    #     m_col = sheet.max_column
    #     # print(s_name, m_row,m_col)
    # # loop
    d_list_list = []
    for shit_m in tqdm(shit_n):
        ws = wb_help[shit_m]
        df_l = pd.read_excel(name_for_save, engine='openpyxl', sheet_name=shit_m)
        print(ws)
        for r in dataframe_to_rows(df_l, index=False, header=False):
            ws.append(r)
        ws.auto_filter.ref = "A6:" + get_column_letter(ws.max_column) \
                             + str(ws.max_row)
        if shit_m == 'D_list':
            for index_x in df_l['tiker']:
                d_list_list.append(index_x)
        if shit_m == 'SPB':
            df_l_spb = df_l.copy()
    if len(d_list_list) !=0:
        my_filter = df_l_spb['tiker'].isin([*d_list_list])
        ws = wb_help['SPB']
        for index_x in df_l_spb[my_filter].index:
            ws[f'B{index_x + 7}'].fill = redFill
        print(f"SPB tiker from d_list -- fill [{len(d_list_list)}]")

        # ws.freeze_panes = 'E7'
    # print(f'Now remove old file [{os.pardir} {name_for_save}]')
    # save_log(linux_path, str(f'Now remove old file [{os.pardir} {name_for_save}]'))
    # os.remove(name_for_save)
    for s_name in shit_n:
        sheet = wb_help[s_name]
        m_row = sheet.max_row
        m_col = sheet.max_column
        print(s_name, m_row, m_col)

    cur_date_for_file_name = str(date.today().day) + '-' + str(date.today().month) + '-' + str(date.today().year)
    name_for_save = str(linux_path) + "d-отчет-" + cur_date_for_file_name + ".xlsx"
    wb_help.save(name_for_save)
    save_log(linux_path, str(wb_help.sheetnames))
    wb_help.save(str(history_path) + "d-отчет-" + cur_date_for_file_name + ".xlsx")  # save to Yandex drive
    pfd = wb_help['D_list']
    wb_help.remove(pfd)
    pfd = wb_help["Z_list"]
    wb_help.remove(pfd)
    save_log(linux_path, str(wb_help.sheetnames))
    name_for_save_crop = str(history_path) + "all-отчет-" + cur_date_for_file_name + ".xlsx"
    wb_help.save(name_for_save_crop)

    print('[', datetime.today(), ']', 'EXCEL format.. [Complite]')
    print('[', datetime.today(), ']', 'EXCEL save to YD.. [Complite]')

    return name_for_save, name_for_save_crop

def pd_df_to_sql(df, db_connection_str, st_name, ind,
                 len_max):  # запись исторических значений при update с investpy  в базу MYSQL
    engine = create_engine(db_connection_str)
    try:
        df.to_sql(name='hist_data', con=engine, if_exists='append')  # append , replace
        # print(st_name,"save_to MYSQL...... [OK]")
    except:
        print('pd_to_sql --- [error]')

def send_email(name_for_save, name_for_save_crop):
    """ пробуем отправлять почту
    для отправки отбираются два файла - name_for_save_crop, name_for_save, склеиваются в один пакет 
    и отправляется по списку client_mail_vip
    """
    file = open('secret.pass', 'r', encoding='ISO-8859-1')
    line_list = file.readlines()
    sender, password = line_list[0][:-1], line_list[1][:-1]

    client_mail_vip = ['volemor@yandex.ru', 'azinaidav@mail.ru',
                       'krotar@mail.ru', 'p.trubitsyn7@yandex.ru', 'shanaev_av@mail.ru']#,'fedorov.efrem@gmail.com' ]
    client_mail = ['p.trubitsyn7@yandex.ru']
    server = smtplib.SMTP("smtp.gmail.com", 587)
    server.starttls()
    file_name = [name_for_save_crop, name_for_save]
    client_my = ' '
    try:
        server.login(sender, password)
        msg = MIMEMultipart()
        msg["From"] = sender
        msg["Subject"] = "Новый отчет " + str(datetime.today().strftime("%Y-%m-%d") + " для моих VIP подписчиков")

        for file in file_name:
            filename = os.path.basename(file)
            ftype, encoding = mimetypes.guess_type(file)
            filetype, subtype = ftype.split("/")
            with open(f"{file}", 'rb') as f:
                file = MIMEApplication(f.read(), subtype)
            file.add_header('content-disposition', 'attachment', filename=filename)
            msg.attach(file)

        for cli in tqdm(client_mail_vip):
            msg['To'] = cli
            server.sendmail(sender, cli, msg.as_string())
            time.sleep(2)
            client_my += str(cli) + ' & '

        '''msg['To'] = client
        server.sendmail(sender, client, msg.as_string())
        time.sleep(2)
        client_my += client + ' & '
        msg['To'] = client2
        server.sendmail(sender, client2, msg.as_string())
        client_my += client2
        #time.sleep(2)
        #server.sendmail(sender, client3, msg.as_string())'''

        return "the message was send - " + client_my

    except Exception as _ex:
        return f"{_ex}\n Check your login"

def dmitry_hist_tab(linux_path, *stope_1):
    tab_name = 'history-all.xlsx'
    dmitry_tab = pd.read_excel(str(linux_path + tab_name), sheet_name='SBP',
                               index_col=0)  ### кстати тут кривое название
    print(dmitry_tab.tiker[pd.DataFrame.notna(dmitry_tab.tiker)])
    dmitry_list_spb = dmitry_tab.tiker[pd.DataFrame.notna(dmitry_tab.tiker)]

    exit()

def main():
    print('программа расчета таблицы по стикерам')
    start_timer = datetime.today()
    print(f'[{start_timer}] START')
    # My constant list
    col_list = my_start()
    db_connection_str = 'mysql+pymysql://python:python@192.168.0.118/hist_data'
    db_connection_hist_bks = 'mysql+pymysql://python:python@192.168.0.118/history_bks'  # создана база данных history_bks - для сохранения  и работы со сделками,,, надо сделать таблицу, для загрузки итоговой сводной таблицы
    if os.name == 'nt':
        linux_path = ''
        history_path = 'D:\\YandexDisk\\корень\\отчеты\\'
    else:
        linux_path = '/mnt/1T/opt/gig/My_Python/st_US/'
        history_path = '/mnt/1T/opt/gig/My_Python/st_US/SAVE/'
    ''' end constant list '''
    save_log(linux_path, ' start normal')
    # TEST MODUL

    # dmitry_hist_tab(linux_path,1)

    # кстати выявлено, что в базе данных есть тикеры , есть они в my_st_list, но в итоговой базе они пропадают.
    # надо придумать способ проверять входные и выходные данные, и какой то делать отчет о целостности.
    # может сваять все в одну таблицу - имя, тикер, отрасль, валюта, рынок, последняя дата загрузки, начальная даты загрузки, номер строчки в итоговой таблице в каждой закладке.
    # save_log(linux_path, ' test_start')
    # input_tables_append(db_connection_str, linux_path ) # модуль для объединения вводных таблиц
    # save_log(linux_path, ' test_stop')
    # exit()
    # year_live_stock() # модуля для подсчета годового роста за 12-- месяцев.
    # End TEST MODUL
    # pd_df_to_sql(big_df_table)

    # history_updater(linux_path, db_connection_str)  # запускаем загрузку и обновление sql базы делает --- отдельный скрипт
    name_for_save = sql_base_make(linux_path, db_connection_str, col_list)  # делаем расчеты для заполнения таблицы
    name_for_save, name_for_save_crop = excel_format(linux_path, name_for_save, history_path)  # форматируем таблицу

    print(send_email(name_for_save, name_for_save_crop))  # отправляем по почте
    # remove old file after sendmail
    # os.remove(name_for_save)
    # os.remove(name_for_save_crop)
    # save_log(linux_path, str(f'Now remove old file [{os.pardir} {name_for_save}]'))
    # save_log(linux_path, str(f'Now remove old file [{os.pardir} {name_for_save_crop}]'))
    message = " all Calculating for [" + str((datetime.today() - start_timer).seconds) + '] sec'
    print(datetime.today(), message)
    save_log(linux_path, message)
    exit()

    # дальше пошло старье....

    # exit()
    # stock_name_table(linux_path) # склеиваем все тикеры в одну таблицу  stock_name для базы данных - и кладем в  .csv  .xlsx
    # history_data(linux_path)# загружаем тикеры из таблицы с market_id и загружаем в таблицу stock_hist_data


if __name__ == "__main__":
    main()

    # my mysql request
    ## ALTER IGNORE TABLE hist_data ADD UNIQUE ( Date, st_id(6));  ## удаляем дубликаты в mysql
    # SELECT * FROM hist_data WHERE st_id = 'A';  ## выделяем строки в столбце, где st_id = A
    ## SELECT * FROM hist_data WHERE st_id = 'A' AND open > 70;
    ## SELECT * FROM hist_data WHERE st_id = 'A' AND open > 70 AND date > '2021-08-09';
    # SELECT Date, Low, High, st_id FROM hist_data WHERE st_id = 'APA' AND date > '2021-08-09';
    ##  SELECT Date, Low, High FROM hist_data WHERE st_id = 'APA' AND date BETWEEN '2021-08-09' AND '2021-09-10';
    # """"
    # пробуем получить список тикеров с максимальными датами
    # df = pd.read_sql('SELECT'+ '*' + 'FROM hist_data', con=db_connection)
    #     ##Select st_id, max(date) as date_max from hist_data group by st_id;
    #
    #
    #     delete FROM hist_data WHERE st_id is null;
    #
    #
    # """
    # my mysql request end###

# CREATE TABLE tiker_branch
# (
# ind INT,
# st_id VARCHAR(6) NOT NULL,
# name VARCHAR (70) NOT NULL,
# branch VARCHAR (50) NOT NULL,
# Currency VARCHAR(3) NOT NULL,
# market VARCHAR(3) NOT NULL
# );
